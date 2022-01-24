import openpyxl

dit_file = openpyxl.load_workbook('dit.xlsx')
esn_list = dit_file["Sheet1"]

esn_per_record = {}

for esn_row in range(2, esn_list.max_row + 1):
    esn = esn_list.cell(esn_row, 2).value
    if esn in esn_per_record:
        current_num_esn = esn_per_record.get(esn);
        esn_per_record[esn] = current_num_esn + 1;
    else:
        esn_per_record[esn] = 1

print(esn_per_record)